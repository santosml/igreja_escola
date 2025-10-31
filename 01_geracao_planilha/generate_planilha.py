"""
Ferramenta para gerar fichas de chamada mensais a partir da planilha base.

O comportamento replica as transformações que aplicamos manualmente:
 - Atualiza o rótulo "Mês:" no cabeçalho de cada aba.
 - Substitui/insere as colunas de chamada com todos os domingos do mês alvo.
 - Atualiza as seções "ASSUNTO DAS AULAS:" e "VISITAS:" com as mesmas datas.
 - Calcula automaticamente os aniversariantes do mês por turma.
 - Replica estilos (largura, cores, bordas) quando novas colunas são criadas.
 - Remove abas que comecem com "Cópia" no arquivo final.

Configuração:
 - O caminho do arquivo de origem, o ano e o mês de destino (numérico) são
   lidos de um arquivo JSON simples (por padrão `config.json` no diretório atual).
 - O arquivo gerado é salvo como `<nome da planilha> - <Nome do mês>.xlsx`
   no diretório indicado na configuração (ou o diretório do arquivo de origem).
"""

from __future__ import annotations

import argparse
import calendar
import json
import re
import unicodedata
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

PORTUGUESE_MONTHS: Dict[int, str] = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}

HEADER_ROW_INDEX = 5  # Linha onde ficam os cabeçalhos de aluno/presença.


@dataclass
class Config:
    source_file: Path
    target_year: int
    target_month: int
    output_directory: Path


def load_config(path: Path) -> Config:
    with path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)

    try:
        source_raw = Path(data["source_file"])
        year = int(data["target_year"])
        month = int(data["target_month"])
    except KeyError as exc:
        raise ValueError(f"Configuração incompleta: campo ausente {exc}.") from exc
    except (TypeError, ValueError) as exc:
        raise ValueError("Ano e mês precisam ser numéricos.") from exc

    if month not in PORTUGUESE_MONTHS:
        raise ValueError("Mês deve estar entre 1 e 12.")

    base_dir = path.parent
    source = (base_dir / source_raw).resolve()

    output_raw = Path(data.get("output_directory", "."))
    output_dir = (base_dir / output_raw).resolve()

    return Config(source_file=source, target_year=year, target_month=month, output_directory=output_dir)


def normalize(text: Optional[str]) -> str:
    if not isinstance(text, str):
        return ""
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii").lower()


def month_sundays(year: int, month: int) -> List[datetime]:
    last_day = calendar.monthrange(year, month)[1]
    sundays: List[datetime] = []
    for day in range(1, last_day + 1):
        current = date(year, month, day)
        if current.weekday() == calendar.SUNDAY:
            sundays.append(datetime.combine(current, datetime.min.time()))
    return sundays


def parse_birth(value) -> Optional[Tuple[int, int]]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.day, value.month
    if isinstance(value, date):
        return value.day, value.month
    if isinstance(value, str):
        text = value.strip()
        if not text or text in {"-", "--"}:
            return None
        nums = re.findall(r"\d+", text)
        if len(nums) >= 2:
            day = int(nums[0])
            month = int(nums[1])
            if 1 <= day <= 31 and 1 <= month <= 12:
                return day, month
    return None


def find_header_columns(ws) -> Tuple[Optional[int], Optional[int]]:
    name_col = None
    birth_col = None
    for cell in ws[HEADER_ROW_INDEX]:
        if isinstance(cell.value, str):
            norm = normalize(cell.value)
            if name_col is None and "nome" in norm:
                name_col = cell.column
            if birth_col is None and "nasc" in norm:
                birth_col = cell.column
    return name_col, birth_col


def collect_students(ws, name_col: int, birth_col: int) -> List[Tuple[str, Optional[Tuple[int, int]]]]:
    students: List[Tuple[str, Optional[Tuple[int, int]]]] = []
    empty_streak = 0
    row_idx = HEADER_ROW_INDEX + 1
    max_row = ws.max_row

    while row_idx <= max_row:
        raw = ws.cell(row=row_idx, column=name_col).value
        if isinstance(raw, str):
            name = raw.strip()
            if not name:
                empty_streak += 1
                if empty_streak >= 3:
                    break
                row_idx += 1
                continue
            normalized = normalize(name)
            if normalized.endswith(":") or any(token in normalized for token in ("present", "ausent", "total", "visit", "assunto", "anivers")):
                break
            empty_streak = 0
            birth = ws.cell(row=row_idx, column=birth_col).value
            students.append((name, parse_birth(birth)))
        else:
            empty_streak += 1
            if empty_streak >= 3:
                break
        row_idx += 1
    return students


def ensure_column_style(ws, src_col: int, dest_col: int) -> None:
    src_letter = get_column_letter(src_col)
    dest_letter = get_column_letter(dest_col)
    src_dim = ws.column_dimensions.get(src_letter)
    if src_dim and src_dim.width is not None:
        ws.column_dimensions[dest_letter].width = src_dim.width
    else:
        # Mantém largura padrão se não houver personalização.
        ws.column_dimensions[dest_letter].width = ws.column_dimensions.get(dest_letter, ws.column_dimensions[dest_letter]).width
    for row in range(1, ws.max_row + 1):
        src_cell = ws.cell(row=row, column=src_col)
        dest_cell = ws.cell(row=row, column=dest_col)
        dest_cell._style = copy(src_cell._style)


def ensure_date_columns(ws, sundays: Sequence[datetime]) -> List[int]:
    existing_cols = sorted([cell.column for cell in ws[HEADER_ROW_INDEX] if isinstance(cell.value, datetime)])
    if not existing_cols:
        return []
    first_col = existing_cols[0]
    needed_cols = [first_col + idx for idx in range(len(sundays))]

    # Insert missing columns to ensure contiguous block.
    for col in needed_cols:
        if col not in existing_cols:
            ws.insert_cols(col)
            # adjust existing indices after insertion
            existing_cols = [c + 1 if c >= col else c for c in existing_cols]
            existing_cols.append(col)
            existing_cols.sort()
            ensure_column_style(ws, col - 1, col)

    # If we still have fewer columns than Sundays (e.g., original planilha tinha menos colunas),
    # append new columns à direita copiando estilo da coluna anterior.
    while len(needed_cols) > len(existing_cols):
        insert_col = existing_cols[-1] + 1
        ws.insert_cols(insert_col)
        ensure_column_style(ws, insert_col - 1, insert_col)
        existing_cols.append(insert_col)

    # Limita ao bloco de interesse.
    return needed_cols


def clear_extra_header_columns(ws, target_cols: Sequence[int]) -> None:
    max_col = max(target_cols)
    col_idx = max_col + 1
    while True:
        cell = ws.cell(row=HEADER_ROW_INDEX, column=col_idx)
        if cell.value in (None, ""):
            break
        cell.value = ""
        col_idx += 1


def update_header_dates(ws, sundays: Sequence[datetime]) -> List[int]:
    target_cols = ensure_date_columns(ws, sundays)
    if not target_cols:
        return []
    for idx, col in enumerate(target_cols):
        cell = ws.cell(row=HEADER_ROW_INDEX, column=col)
        cell.value = sundays[idx]
        cell.number_format = "dd/mm"
    clear_extra_header_columns(ws, target_cols)

    # Garante que as células abaixo existam e tenham estilo.
    last_col = target_cols[-1]
    prev_col = target_cols[-2] if len(target_cols) >= 2 else last_col
    ensure_column_style(ws, prev_col, last_col)
    return target_cols


def find_label_cell(ws, label: str):
    label_norm = normalize(label)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, str) and normalize(cell.value) == label_norm:
                return cell
    return None


def update_section_dates(ws, label: str, sundays: Sequence[datetime]) -> None:
    cell = find_label_cell(ws, label)
    if not cell:
        return
    label_row = cell.row
    label_col = cell.column
    next_row = label_row + 1
    while next_row <= ws.max_row:
        val = ws.cell(row=next_row, column=label_col).value
        if isinstance(val, str):
            normalized_val = normalize(val)
            if normalized_val.endswith(":") and normalized_val != normalize(label):
                break
        next_row += 1
    needed = len(sundays)
    available = next_row - label_row - 1
    if available < needed:
        ws.insert_rows(next_row, amount=needed - available)
        next_row += needed - available
    for idx, dt in enumerate(sundays):
        target_cell = ws.cell(row=label_row + 1 + idx, column=label_col)
        target_cell.value = dt
        target_cell.number_format = "dd/mm"
    for cleanup_row in range(label_row + 1 + len(sundays), next_row):
        ws.cell(row=cleanup_row, column=label_col).value = ""


def update_birthdays(ws, students: Sequence[Tuple[str, Optional[Tuple[int, int]]]], target_month: int) -> None:
    label_cell = find_label_cell(ws, "ANIVERSARIANTES:")
    if not label_cell:
        return
    label_row = label_cell.row
    label_col = label_cell.column
    next_row = label_row + 1
    while next_row <= ws.max_row:
        val = ws.cell(row=next_row, column=label_col).value
        if isinstance(val, str):
            normalized = normalize(val)
            if normalized.endswith(":") and normalized != "aniversariantes:":
                break
        next_row += 1
    available = max(0, next_row - label_row - 1)
    birthday_entries: List[Tuple[int, str, str]] = []
    for name, parsed in students:
        if not parsed:
            continue
        day, month = parsed
        if month == target_month:
            sort_key = normalize(name)
            birthday_entries.append((day, sort_key, name))

    birthday_entries.sort(key=lambda item: (item[0], item[1]))
    birthdays: List[str] = [f"{name} - {day:02d}/{target_month:02d}" for day, _key, name in birthday_entries]
    if len(birthdays) > available:
        ws.insert_rows(next_row, amount=len(birthdays) - available)
        next_row += len(birthdays) - available
    for idx, text in enumerate(birthdays):
        ws.cell(row=label_row + 1 + idx, column=label_col, value=text)
    for cleanup_row in range(label_row + 1 + len(birthdays), next_row):
        ws.cell(row=cleanup_row, column=label_col).value = ""


def update_month_label(ws, month_name: str) -> None:
    label = f"M\u00EAs: {month_name}"
    for row in ws.iter_rows(min_row=1, max_row=3):
        for cell in row:
            if isinstance(cell.value, str) and "mes" in normalize(cell.value):
                cell.value = label


def process_sheet(ws, sundays: Sequence[datetime], month_name: str, target_month: int) -> None:
    update_month_label(ws, month_name)
    date_cols = update_header_dates(ws, sundays)
    if not date_cols:
        return
    name_col, birth_col = find_header_columns(ws)
    if name_col is None or birth_col is None:
        return
    students = collect_students(ws, name_col, birth_col)
    update_section_dates(ws, "ASSUNTO DAS AULAS:", sundays)
    update_section_dates(ws, "VISITAS:", sundays)
    update_birthdays(ws, students, target_month)


def generate_planilha(config: Config) -> Path:
    if not config.source_file.exists():
        raise FileNotFoundError(f"Arquivo de origem não encontrado: {config.source_file}")

    wb = openpyxl.load_workbook(config.source_file)
    month_name = PORTUGUESE_MONTHS[config.target_month]
    sundays = month_sundays(config.target_year, config.target_month)

    for ws in list(wb.worksheets):
        if normalize(ws.title).startswith("copia"):
            wb.remove(ws)
            continue
        process_sheet(ws, sundays, month_name, config.target_month)

    output_dir = config.output_directory or config.source_file.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    output_name = f"{config.source_file.stem} - {month_name}.xlsx"
    output_path = output_dir / output_name
    wb.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Gera fichas mensais da EBD a partir de uma planilha base.")
    parser.add_argument(
        "--config",
        type=Path,
        default=Path("config.json"),
        help="Arquivo JSON com parâmetros (padrão: config.json).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    config = load_config(args.config)
    output_path = generate_planilha(config)
    print(f"Arquivo gerado: {output_path}")


if __name__ == "__main__":
    main()
